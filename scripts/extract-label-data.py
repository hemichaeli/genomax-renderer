"""
GenoMAX² — Production Label Data Extractor
============================================
Reads both Excel files and outputs structured JSON for all 168 SKUs.
Each SKU gets a complete label data object mapped to the 7-zone system.

Usage: python scripts/extract-label-data.py
Output: design-system/data/production-labels-maximo.json
        design-system/data/production-labels-maxima.json
"""

import json
import os
import openpyxl

EXCEL_FILES = {
    "maximo": r"G:\My Drive\Work\GenoMAX²\Design\Lables\V6 (1)\GENOMAX_MAXimo_LABEL_READY_v2.xlsx",
    "maxima": r"G:\My Drive\Work\GenoMAX²\Design\Lables\V6 (1)\GENOMAX_MAXima_LABEL_READY_v2.xlsx",
}

ACCENT_COLORS = {
    "MAXimo²": "#7A1E2E",
    "MAXima²": "#7A304A",
}

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "design-system", "data")

# Column indices (0-based) from Excel
COL = {
    "module_id": 0,
    "module_code": 1,
    "product_line": 2,
    "os": 3,
    "biological_system": 4,
    "function_name": 5,
    "ingredient_descriptor": 6,
    "supliful_handle": 7,
    "status": 8,
    "ingredient_name_label": 9,
    "net_quantity": 10,
    "layer": 11,
    "suggested_use": 12,
    "dosing_window": 13,
    "safety_notes": 14,
    "contraindications": 15,
    "fda_disclaimer": 16,
    "front_label_text": 17,
    "back_label_text": 18,
    "product_form": 19,
    "container_type": 20,
    "label_format": 21,
    "label_w_in": 22,
    "label_h_in": 23,
    "label_w_mm": 24,
    "label_h_mm": 25,
    "bleed_in": 26,
    "safe_zone_in": 27,
    "print_w_in": 28,
    "print_h_in": 29,
}


def safe_str(val):
    """Convert value to string safely."""
    if val is None:
        return ""
    return str(val).strip()


def build_label_data(row, os_name):
    """Build a complete label data object for one SKU."""
    accent = ACCENT_COLORS.get(os_name, "#7A1E2E")
    module_code = safe_str(row[COL["module_code"]])
    ingredient_name = safe_str(row[COL["ingredient_name_label"]])
    bio_system = safe_str(row[COL["biological_system"]])
    function_name = safe_str(row[COL["function_name"]])
    product_form = safe_str(row[COL["product_form"]])
    container = safe_str(row[COL["container_type"]])
    label_format = safe_str(row[COL["label_format"]])

    return {
        "_meta": {
            "module_id": safe_str(row[COL["module_id"]]),
            "module_code": module_code,
            "os": os_name,
            "product_line": safe_str(row[COL["product_line"]]),
            "supliful_handle": safe_str(row[COL["supliful_handle"]]),
            "status": safe_str(row[COL["status"]]),
        },
        "format": {
            "container_type": container,
            "label_format": label_format,
            "product_form": product_form,
            "dimensions": {
                "label_w_in": row[COL["label_w_in"]],
                "label_h_in": row[COL["label_h_in"]],
                "label_w_mm": row[COL["label_w_mm"]],
                "label_h_mm": row[COL["label_h_mm"]],
                "bleed_in": row[COL["bleed_in"]],
                "safe_zone_in": row[COL["safe_zone_in"]],
                "print_w_in": row[COL["print_w_in"]],
                "print_h_in": row[COL["print_h_in"]],
            },
        },
        "front_panel": {
            "zone_1": {
                "brand_name": "GenoMAX²",
                "module_code": module_code,
            },
            "zone_2": {
                "text": "BIOLOGICAL OS MODULE",
            },
            "zone_3": {
                "ingredient_name": ingredient_name.upper(),
            },
            "zone_4": {
                "descriptor": safe_str(row[COL["ingredient_descriptor"]]),
                "biological_system": f"{bio_system.upper()} · {function_name.upper()}",
            },
            "zone_5": {
                "variant_name": os_name,
                "accent_color": accent,
            },
            "zone_6": {
                "type": {"label": "TYPE", "value": product_form},
                "function": {"label": "FUNCTION", "value": function_name},
                "status": {"label": "STATUS", "value": "Active" if safe_str(row[COL["status"]]) == "VALID" else safe_str(row[COL["status"]])},
            },
            "zone_7": {
                "version_info": f"v1.0 · {module_code} · Clinical Grade",
                "net_quantity": f"DIETARY SUPPLEMENT · {safe_str(row[COL['net_quantity']])}",
            },
        },
        "back_panel": {
            "suggested_use": safe_str(row[COL["suggested_use"]]),
            "safety_notes": safe_str(row[COL["safety_notes"]]),
            "contraindications": safe_str(row[COL["contraindications"]]),
            "fda_disclaimer": safe_str(row[COL["fda_disclaimer"]]),
            "layer": safe_str(row[COL["layer"]]),
            "front_label_text": safe_str(row[COL["front_label_text"]]),
            "back_label_text": safe_str(row[COL["back_label_text"]]),
        },
        "brand_signature": {
            "ceiling_color": accent,
            "ceiling_height": "2px",
            "brand_tracking": "0.18em",
            "brand_rule_opacity": 0.25,
        },
    }


def extract_all():
    """Extract all SKUs from both Excel files."""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    summary = {}

    for system_name, filepath in EXCEL_FILES.items():
        print(f"\n{'='*60}")
        print(f"Processing: {system_name}")
        print(f"File: {filepath}")

        wb = openpyxl.load_workbook(filepath, read_only=True)
        # Get the first sheet (MAXimo² or MAXima²)
        sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]

        os_name = "MAXimo²" if system_name == "maximo" else "MAXima²"
        skus = []
        format_counts = {}

        for i, row_data in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            if row_data[0] is None:
                continue

            label_data = build_label_data(row_data, os_name)
            skus.append(label_data)

            # Count by format
            fmt = label_data["format"]["label_format"]
            format_counts[fmt] = format_counts.get(fmt, 0) + 1

        # Write output
        output_file = os.path.join(OUTPUT_DIR, f"production-labels-{system_name}.json")
        with open(output_file, "w", encoding="utf-8") as f:
            json.dump({
                "_meta": {
                    "system": f"GenoMAX² {os_name}",
                    "version": "3.0-LOCKED",
                    "total_skus": len(skus),
                    "format_breakdown": format_counts,
                    "source_file": os.path.basename(filepath),
                    "generated": "2026-04-07",
                },
                "skus": skus,
            }, f, indent=2, ensure_ascii=False)

        print(f"  SKUs extracted: {len(skus)}")
        print(f"  Formats: {format_counts}")
        print(f"  Output: {output_file}")
        summary[system_name] = {"count": len(skus), "formats": format_counts}

        wb.close()

    # Write combined summary
    summary_file = os.path.join(OUTPUT_DIR, "production-summary.json")
    with open(summary_file, "w", encoding="utf-8") as f:
        json.dump({
            "total_skus": sum(s["count"] for s in summary.values()),
            "systems": summary,
            "version": "3.0-LOCKED",
        }, f, indent=2)

    print(f"\n{'='*60}")
    print(f"TOTAL: {sum(s['count'] for s in summary.values())} SKUs extracted")
    print(f"Summary: {summary_file}")


if __name__ == "__main__":
    extract_all()
